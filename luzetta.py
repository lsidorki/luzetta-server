import html
import re
import time

import tidalapi
import argparse
import openpyxl
import requests
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape
from bs4 import BeautifulSoup


class DataEntry:
    autor_wpisu: ""
    category: ""
    composer: ""
    lyricist: ""
    artist: ""
    title: ""
    album: ""
    label: ""


class CreditsObject:
    def __init__(self, artist, title, album, composer, lyricist, label, year):
        self.artist = artist
        self.title = title
        self.album = album
        self.composer = composer
        self.lyricist = lyricist
        self.label = label
        self.year = year

    def __str__(self):
        return (f"Artist: {self.artist}\n"
                f"Title: {self.title}\n"
                f"Album: {self.album}\n"
                f"Composer: {self.composer}\n"
                f"Lyricist: {self.lyricist}\n"
                f"Label: {self.label}\n"
                f"Year: {self.year}")


album_credits_dict = dict([])


composer_dict = {'Composer', 'Producer', 'Co-Producer', 'Misc. Prod.'
                 'Featured Artist', 'Vocals', 'Associated Performer',
                 'Beat Boxing', 'Background Vocal',
                 'Drums', 'Electric Guitar', 'Lead Guitar', 'Percussion', 'Piano', 'Guitar', 'Synthesizer', 'Bass',
                 'Saxophone', 'Upright Bass', 'Viola da Gamba', 'Keyboards', 'Additional Synthesizer',
                 'Bass guitar', 'Backing Vocals', 'Drum Programming', 'Violin', 'Viola', 'Bass Trombone',
                 'Trombone', 'Trumpet', 'Cello', 'All Instruments', 'Drum Programmer', 'Programmer',
                 'Instrumentation', 'Drum Machine', 'Organ', 'Mellotron', 'Flute', 'Clavichord', 'Clarinet',
                 'Wurlitzer Electric Piano', 'Saxophones', 'Horn'}


lyricist_dict = {'Lyricist', 'Writer', 'Lead Vocalist', 'Co-Writer', 'Lead Vocals', 'Arranger', 'Vocal',
                 'Songwriter', 'Vocalist', 'Author'}


# gs_url = "http://156.17.147.20/GSImportExportService/GSImportExportService.asmx"
gs_url = "http://10.100.64.212/GSImportExportService/GSImportExportService.asmx"
# gs_url = "http://156.17.39.69/GSImportExportService/GSImportExportService.asmx"


def get_parser(h):
    parser = argparse.ArgumentParser(description='LUZetta tracklist processor.', add_help=h)
    parser.add_argument('-i', '--input', nargs=1, metavar='INPUT_FILE_PATH', help='Excel file path', required=False)
    parser.add_argument('-o', '--output', nargs=1, metavar='OUTPUT_PATH', help='Output file path', required=False)
    parser.add_argument('-f', '--filter', nargs=1, metavar='FILTER_STRING', help='Filter by author', required=False)

    return parser


def init_tidal_session():
    session = tidalapi.Session()
    # Will run until you visit the printed url and link your account
    session.login_oauth_simple()

    return session


def get_track_info(result):
    for track in result['tracks']:
        return track


def fetch_track_credits(track_name, album_tracklist):
    for single_track in album_tracklist:
        if single_track['item']['title'] == track_name:
            return single_track


def exists_in_cache(album_key):
    return album_credits_dict.get(album_key)


def add_to_cache(album_key, credits_value):
    album_credits_dict[album_key] = credits_value


def fetch_tidal_data(session, query):
    composer = set()
    lyricist = set()
    missing = set()

    result = session.search(query=query, models=None, limit=50, offset=0)
    track_info = get_track_info(result)

    if track_info is not None:
        album_copyright = None
        if exists_in_cache(track_info.artist.name + " - " + track_info.album.name) is not None:
            album_credits = album_credits_dict[track_info.artist.name + " - " + track_info.album.name]
        else:
            album_info = session.album(track_info.album.id)
            album_credits = album_info.credits()
            album_copyright = album_info.copyright
            add_to_cache(track_info.artist.name + " - " + track_info.album.name, album_credits)
        track_credits = fetch_track_credits(track_info.name, album_credits)

        for credit in track_credits['credits']:
            for contributor in credit['contributors']:
                if credit['type'] in composer_dict:
                    composer.add(contributor['name'])
                elif credit['type'] in lyricist_dict:
                    lyricist.add(contributor['name'])
                else:
                    missing.add(credit['type'])

        if len(composer) == 0:
            composer.add(track_info.artist.name)
        if len(lyricist) == 0:
            lyricist.add(track_info.artist.name)

        return CreditsObject(track_info.artist.name, track_info.name, track_info.album.name, ', '.join(composer),
                             ', '.join(lyricist), album_copyright, str(track_info.album.year))


def process_input_data(input_data):
    workbook = openpyxl.load_workbook(input_data)
    sheet = workbook.active
    row = sheet.max_row
    column = sheet.max_column

    data_list = []

    autor_wpisu = 0
    category = 0
    composer = 0
    lyricist = 0
    artist = 0
    title = 0
    album = 0
    label = 0

    # Read and determine headers
    for i in range(1, column + 1):
        cell = sheet.cell(row=1, column=i)
        if cell.value == "Autor_wpisu":
            autor_wpisu = i
        elif cell.value == "Category":
            category = i
        elif cell.value == "Composer":
            composer = i
        elif cell.value == "Writer/Lyricist":
            lyricist = i
        elif cell.value == "Artist":
            artist = i
        elif cell.value == "Title":
            title = i
        elif cell.value == "Album":
            album = i
        elif cell.value == "Label":
            label = i

    # Read all the data
    for i in range(2, row + 1):
        data_entry = DataEntry()
        data_entry.autor_wpisu = get_sheet_value(sheet, i, autor_wpisu)
        data_entry.category = get_sheet_value(sheet, i, category)
        data_entry.composer = get_sheet_value(sheet, i, composer)
        data_entry.lyricist = get_sheet_value(sheet, i, lyricist)
        data_entry.artist = get_sheet_value(sheet, i, artist)
        data_entry.title = get_sheet_value(sheet, i, title)
        data_entry.album = get_sheet_value(sheet, i, album)
        data_entry.label = get_sheet_value(sheet, i, label)
        data_list.append(data_entry)

    print("Processed successfully: " + str(len(data_list)) + " entries")

    return data_list


def get_sheet_value(sheet, row, column):
    cell = sheet.cell(row=row, column=column)
    return cell.value


def update_song_album(song, album_title, namespaces):
    album_list = song.findall('gs_s:Album', namespaces)
    for album in album_list:
        if album.get("name") is None or album.get("name") == "" or album.get("name") == "-":
            album.set("name", album_title)


def update_song_label(song, label, namespaces):
    if label is None or label == "":
        pass

    additional_list = song.findall('gs_sl:Additional', namespaces)
    for additional in additional_list:
        if additional.get("label") is None or additional.get("label") == "" or additional.get("label") == "-":
            additional.set("label", label)


def update_song_composer(participants, composer, namespaces):
    if composer is None or composer == "":
        pass

    composers_list = participants.findall('gs_s:Composers', namespaces)
    if len(composers_list) == 0:
        composers_node = ET.Element('gs_s:Composers')
        composer = composer[:100] if len(composer) > 100 else composer
        composers_node.set("name", composer)
        participants.append(composers_node)
    else:
        for composers in composers_list:
            if composers.get("name") is None or composers.get("name") == "" or composers.get("name") == "-":
                composers.set("name", composer)


def update_song_lyricist(participants, lyricist, namespaces):
    if lyricist is None or lyricist == "":
        pass

    lyricist_list = participants.findall('gs_s:Lyricist', namespaces)
    if len(lyricist_list) == 0:
        lyricist_node = ET.Element('gs_s:Lyricist')
        lyricist = lyricist[:100] if len(lyricist) > 100 else lyricist
        lyricist_node.set("name", lyricist)
        participants.append(lyricist_node)
    else:
        for lyricist in lyricist_list:
            if lyricist.get("name") is None or lyricist.get("name") == "" or lyricist.get("name") == "-":
                lyricist.set("name", lyricist)


def update_song_participants(song, composer, lyricist, namespaces):
    song_codes_list = song.findall('gs_s:SongCodes', namespaces)
    if len(song_codes_list) == 0:
        song_codes_element = ET.Element('gs_s:SongCodes')
        participants_element = ET.Element('gs_s:Participants')
        song_codes_element.append(participants_element)
        song.append(song_codes_element)
        song_codes_list = song.findall('gs_s:SongCodes')
        for song_code in song_codes_list:
            participants_list = song_code.findall('gs_s:Participants')
            for participants in participants_list:
                update_song_composer(participants, composer, namespaces)
                update_song_lyricist(participants, lyricist, namespaces)
        pass

    for song_code in song_codes_list:
        participants_list = song_code.findall('gs_s:Participants', namespaces)
        for participants in participants_list:
            update_song_composer(participants, composer, namespaces)
            update_song_lyricist(participants, lyricist, namespaces)


def update_song_xml_data(root, namespaces, credits_data):
    result = root.findall('gs_s:Song', namespaces)
    for song in result:
        update_song_album(song, credits_data.album, namespaces)
        update_song_label(song, credits_data.label, namespaces)
        update_song_participants(song, credits_data.composer, credits_data.lyricist, namespaces)
    return root


def fetch_gselector_data(credits_data, entry_artist, entry_title):
    payload_template = """<?xml version="1.0" encoding="utf-8"?>
                <soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
                  <soap:Header>
                    <ConnectionHeader xmlns="http://www.rcsworks.com/webservices/gselector/" />
                  </soap:Header>
                  <soap:Body>
                    <FindSong xmlns="http://www.rcsworks.com/webservices/gselector/">
                      <title>{title}</title>
                      <artist>{artist}</artist>
                    </FindSong>
                  </soap:Body>
                </soap:Envelope>"""
    payload = payload_template.format(title=html.escape(entry_title), artist=html.escape(entry_artist))
    headers = {
        'Content-Type': 'text/xml; charset=utf-8',
        'SOAPAction': 'http://www.rcsworks.com/webservices/gselector/FindSong'
    }
    response = requests.request("POST", gs_url, headers=headers, data=payload.encode('ascii', 'xmlcharrefreplace'))

    return response


def adjust_selector_data(selector_data, credits_data):
    xml = BeautifulSoup(selector_data.text, 'xml')
    namespaces = {'gs_s': 'SongSchemaGS',
                  'gs_pe': 'PEContentSchemaGS',
                  'gs_sl': 'SLContentSchemaGS',
                  'gs_err': 'OperationStatusSchemaGS'}

    root = ET.fromstring(xml.get_text())
    song_error = root.findall('gs_err:SongError', namespaces)
    if len(song_error) > 0:
        error_message = "Error fetching the data for: {artist} - {title}"
        print(error_message.format(title=credits_data.title, artist=credits_data.artist))
        return None

    updated_xml = update_song_xml_data(root, namespaces, credits_data)

    success_message = "Successfully processed the data for: {artist} - {title}"
    print(success_message.format(title=credits_data.title, artist=credits_data.artist))

    return updated_xml


def standardize(title):
    title = re.sub(r'\((feat\..*)\)', '', title)
    return title


def export_gselector_data(selector_data):
    payload_template = """<?xml version="1.0" encoding="utf-8"?>
                        <soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
                          <soap:Header>
                            <ConnectionHeader xmlns="http://www.rcsworks.com/webservices/gselector/" />
                          </soap:Header>
                          <soap:Body>
                            <ImportSongs xmlns="http://www.rcsworks.com/webservices/gselector/">
                              <xmlIn>{xml_input}</xmlIn>
                            </ImportSongs>
                          </soap:Body>
                        </soap:Envelope>"""
    ET.register_namespace('gs_s', 'SongSchemaGS')
    ET.register_namespace('gs_pe', 'PEContentSchemaGS')
    ET.register_namespace('gs_sl', 'SLContentSchemaGS')
    ET.register_namespace('gs_err', 'OperationStatusSchemaGS')
    ET.register_namespace('', 'GSelectorSchemaGS')

    # kodowanie requestu zrobić

    inner_xml = ET.tostring(selector_data, encoding='us-ascii', method='xml').decode("UTF-8")
    payload = payload_template.format(xml_input=escape(inner_xml))
    headers = {
        'Content-Type': 'text/xml; charset=utf-8',
        'SOAPAction': 'http://www.rcsworks.com/webservices/gselector/ImportSongs'
    }
    response = requests.request("POST", gs_url, headers=headers, data=payload.encode('ascii', 'xmlcharrefreplace'))

    return response


def check_export_result(export_response):
    xml = BeautifulSoup(export_response.text, 'xml')
    namespaces = {'gs_s': 'SongSchemaGS',
                  'gs_pe': 'PEContentSchemaGS',
                  'gs_sl': 'SLContentSchemaGS',
                  'gs_err': 'OperationStatusSchemaGS'}

    root = ET.fromstring(xml.get_text())

    if root.get("result") is not None and root.get("result") == "success":
        print("Successfully Exported Data")
    else:
        print("Error Exporting Data")
    pass


def process_selector_entry(fetched_credits, artist, title):
    selector_response = fetch_gselector_data(fetched_credits, entry.artist, entry.title)
    adjusted_selector_data = adjust_selector_data(selector_response, fetched_credits)
    time.sleep(1)
    if adjusted_selector_data is not None:
        export_result = export_gselector_data(adjusted_selector_data)
        check_export_result(export_result)
        time.sleep(1)


if __name__ == '__main__':
    # Define parser
    args_parser = get_parser(h=True)
    args = args_parser.parse_args()
    entries = []
    author_filter = ""

    if args.filter is not None:
        author_filter = args.filter[0]
    if args.input is not None:
        entries = process_input_data(args.input[0])
        tidal_session = init_tidal_session()
        for entry in entries:
            if author_filter == "" or entry.autor_wpisu == author_filter:
                credits_object = fetch_tidal_data(tidal_session, entry.artist.split(" ft. ")[0] + " " + standardize(entry.title))
                try:
                    process_selector_entry(credits_object, entry.artist, entry.title)
                except (TypeError, AttributeError):
                    error_message = "{status} exporting the data for: {artist} - {title}"
                    print(error_message.format(title=entry.title, artist=entry.artist, status='ERROR'))
                    print(error_message.format(title=entry.title, artist=entry.artist, status='RETRY'))
                    try:
                        process_selector_entry(credits_object, entry.artist, entry.title)
                    except (TypeError, AttributeError):
                        print(error_message.format(title=entry.title, artist=entry.artist, status='ERROR'))
                        print(error_message.format(title=entry.title, artist=entry.artist, status='SKIP'))


    print('luzetta-server: Processed successfully.')
